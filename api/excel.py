from http.server import BaseHTTPRequestHandler
import json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

W="FFFFFF"; DARK="1F2937"; GBG="E8F5E9"; RBG="FFEBEE"
GR="1B5E20"; RD="B71C1C"; GD="E65100"; MT="757575"
GOLD="C9A84C"; GOLD_BG="FEF9ED"; HDR_BG="374151"

def fill(color):
    return PatternFill(fill_type="solid", fgColor=color)

def fnt(color="1F2937", bold=False, size=10, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)

def aln(h="left"):
    return Alignment(horizontal=h, vertical="center")

def brd():
    s = Side(style="thin", color="E0E0E0")
    return Border(top=s, bottom=s, left=s, right=s)

def sc(cell, color="1F2937", bg=None, bold=False, size=10, italic=False, h="left"):
    cell.font = fnt(color=color, bold=bold, size=size, italic=italic)
    if bg: cell.fill = fill(bg)
    cell.alignment = aln(h=h)
    cell.border = brd()

def make_kc_sheet(ws, zavodlar, filter_zavod, dan, gacha, davr_label):
    from datetime import datetime
    def parse_d(s):
        try: return datetime.strptime(s, "%d.%m.%Y")
        except: return datetime.min
    def in_davr(sana):
        if not dan and not gacha: return True
        d = parse_d(sana)
        if dan and d < datetime.strptime(dan, "%Y-%m-%d"): return False
        if gacha and d > datetime.strptime(gacha, "%Y-%m-%d"): return False
        return True

    HDR = ["Sana","Zavod","Tur","+/-","Kimga","Kirim(g)","Naqt($)","Kurs($/g)","Naqt->g","Lom(g)","LomKurs","Lom($)","Chiqim(g)","Ostatka(g)"]
    WCOL = [13,13,10,5,14,12,12,11,11,11,10,14,13,13]

    ws.merge_cells("A1:N1")
    ws["A1"] = "TILLA HISOB - Kirdi-Chiqdi" + (" - " + filter_zavod if filter_zavod else " (Barcha)")
    ws["A1"].font = fnt(W, bold=True, size=13)
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = "Davr: " + davr_label
    ws["A2"].font = fnt(MT, size=9, italic=True)
    ws["A2"].fill = fill("F5F5F5")
    ws["A2"].alignment = aln("center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 4

    for ci, (h, w) in enumerate(zip(HDR, WCOL), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(W, bold=True, size=9)
        c.fill = fill(HDR_BG)
        c.alignment = aln("center")
        c.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 22

    all_rows = []
    for z in zavodlar:
        if filter_zavod and z["nom"] != filter_zavod: continue
        for t in z["turlar"]:
            bal = 0.0
            for op in t["tarix"]:
                if op["tip"] == "mol": bal += op["gramm"]
                else: bal = max(0, bal - (op.get("jami") or 0))
                if not in_davr(op["sana"]): continue
                all_rows.append({"sana":op["sana"],"zavod":z["nom"],"tur":t["nom"],"tip":op["tip"],"op":op,"ostatka":round(bal,2)})
    all_rows.sort(key=lambda r: parse_d(r["sana"]))

    ri = 5
    for row in all_rows:
        op = row["op"]; is_k = row["tip"] == "mol"
        bg = GBG if is_k else RBG
        vals = [
            row["sana"], row["zavod"], row["tur"],
            "+" if is_k else "-",
            "" if is_k else (op.get("kimga") or ""),
            round(op["gramm"],2) if is_k else None,
            op.get("naqtSumma") if not is_k else None,
            op.get("naqtKurs") if not is_k else None,
            round(op.get("naqtGramm",0),2) if not is_k else None,
            round(op.get("lomGramm",0),2) if not is_k else None,
            op.get("lomKurs") if not is_k else None,
            round(op.get("lomPul",0),2) if not is_k else None,
            round(op.get("jami",0),2) if not is_k else None,
            row["ostatka"]
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill(bg)
            c.border = brd()
            if ci == 4:
                c.font = fnt(GR if is_k else RD, bold=True, size=12)
                c.alignment = aln("center")
            elif ci == 6 and is_k:
                c.font = fnt(GR, bold=True, size=10); c.alignment = aln("right")
            elif ci == 13 and not is_k:
                c.font = fnt(RD, bold=True, size=10); c.alignment = aln("right")
            elif ci == 14:
                c.font = fnt("1A237E", bold=True, size=10); c.alignment = aln("right")
            elif ci in [7,8,9,10,11,12]:
                c.font = fnt("424242", size=10); c.alignment = aln("right")
            else:
                c.font = fnt("212121", size=10); c.alignment = aln("left" if ci<=5 else "right")
        ws.row_dimensions[ri].height = 18
        ri += 1

    ri += 1
    tK = round(sum(r["op"]["gramm"] for r in all_rows if r["tip"]=="mol"), 2)
    tC = round(sum(r["op"].get("jami",0) for r in all_rows if r["tip"]=="tolov"), 2)
    tN = round(sum(r["op"].get("naqtSumma",0) for r in all_rows if r["tip"]=="tolov"), 2)
    tL = round(sum(r["op"].get("lomPul",0) for r in all_rows if r["tip"]=="tolov"), 2)
    fin = {}
    for r in all_rows: fin[r["zavod"]+"|"+r["tur"]] = r["ostatka"]
    tO = round(sum(fin.values()), 2)

    ws.merge_cells(f"A{ri}:E{ri}")
    ws[f"A{ri}"] = "JAMI"
    ws[f"A{ri}"].font = fnt(W, bold=True, size=11)
    ws[f"A{ri}"].fill = fill(DARK)
    ws[f"A{ri}"].alignment = aln("center")
    ws.row_dimensions[ri].height = 24

    pairs = [("Kirim:", f"+{tK:.2f}g", "5AB87A"), ("Chiqim:", f"-{tC:.2f}g", "E05A5A"),
             ("Naqt:", f"{tN:,.2f}$", GD), ("Lom:", f"{tL:,.2f}$", GD)]
    col = 6
    for lbl, val, clr in pairs:
        c = ws.cell(row=ri, column=col, value=lbl)
        c.font = fnt("AAAAAA", size=9); c.fill = fill(DARK); c.alignment = aln("right")
        c = ws.cell(row=ri, column=col+1, value=val)
        c.font = fnt(clr, bold=True, size=10); c.fill = fill(DARK); c.alignment = aln("left")
        col += 2

    ri += 1
    ws.merge_cells(f"A{ri}:K{ri}")
    ws[f"A{ri}"] = "Oxirgi ostatka:"
    ws[f"A{ri}"].font = fnt(MT, size=10, italic=True)
    ws[f"A{ri}"].fill = fill(GOLD_BG)
    ws[f"A{ri}"].alignment = aln("right")
    ws.merge_cells(f"L{ri}:N{ri}")
    ws[f"L{ri}"] = f"{tO:.2f} g"
    ws[f"L{ri}"].font = fnt(GOLD, bold=True, size=16)
    ws[f"L{ri}"].fill = fill(GOLD_BG)
    ws[f"L{ri}"].alignment = aln("center")
    ws.row_dimensions[ri].height = 30
    ws.freeze_panes = "A5"

def make_hisobot_sheet(ws, zavodlar, filter_zavod, dan, gacha, davr_label):
    from datetime import datetime
    def parse_d(s):
        try: return datetime.strptime(s, "%d.%m.%Y")
        except: return datetime.min
    def in_davr(sana):
        if not dan and not gacha: return True
        d = parse_d(sana)
        if dan and d < datetime.strptime(dan, "%Y-%m-%d"): return False
        if gacha and d > datetime.strptime(gacha, "%Y-%m-%d"): return False
        return True

    ws.merge_cells("A1:H1")
    ws["A1"] = "HISOBOT - Tur boyicha kirdi-chiqdi"
    ws["A1"].font = fnt(W, bold=True, size=13)
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Davr: " + davr_label
    ws["A2"].font = fnt(MT, size=9, italic=True)
    ws["A2"].fill = fill("F5F5F5")
    ws["A2"].alignment = aln("center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 4

    hdrs = ["Zavod","Tur","Kirim(g)","Chiqim(g)","Ostatka(g)","Naqt($)","Lom($)","Jami pul($)"]
    wcols = [14,12,12,12,13,14,14,14]
    for ci, (h, w) in enumerate(zip(hdrs, wcols), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(W, bold=True, size=9)
        c.fill = fill(HDR_BG)
        c.alignment = aln("center")
        c.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 22

    ri = 5
    gK=gC=gO=gN=gL = 0
    for z in zavodlar:
        if filter_zavod and z["nom"] != filter_zavod: continue
        for t in z["turlar"]:
            tk=tc=tn=tl=bal = 0
            for op in t["tarix"]:
                if op["tip"]=="mol": bal+=op["gramm"]
                else: bal=max(0,bal-(op.get("jami") or 0))
                if not in_davr(op["sana"]): continue
                if op["tip"]=="mol": tk+=op["gramm"]
                else: tc+=op.get("jami",0); tn+=op.get("naqtSumma",0); tl+=op.get("lomPul",0)
            o = round(bal, 2)
            bg = "FAFAFA" if ri%2==0 else W
            vals = [z["nom"],t["nom"],round(tk,2),round(tc,2),o,round(tn,2),round(tl,2),round(tn+tl,2)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = fill(bg); c.border = brd()
                if ci==3: c.font=fnt(GR,bold=True,size=10); c.alignment=aln("right")
                elif ci==4: c.font=fnt(RD,bold=True,size=10); c.alignment=aln("right")
                elif ci==5: c.font=fnt(GOLD,bold=True,size=11); c.alignment=aln("right")
                elif ci in[6,7,8]: c.font=fnt(GD,size=10); c.alignment=aln("right")
                else: c.font=fnt("212121",size=10); c.alignment=aln("left")
            ws.row_dimensions[ri].height = 20
            ri+=1; gK+=tk; gC+=tc; gO+=o; gN+=tn; gL+=tl

    ri += 1
    ws.merge_cells(f"A{ri}:B{ri}")
    ws[f"A{ri}"] = "JAMI"
    ws[f"A{ri}"].font = fnt(W, bold=True, size=11)
    ws[f"A{ri}"].fill = fill(DARK)
    ws[f"A{ri}"].alignment = aln("center")
    for ci, (v, clr) in enumerate([(None,None),(None,None),(round(gK,2),GR),(round(gC,2),RD),(round(gO,2),GOLD),(round(gN,2),GD),(round(gL,2),GD),(round(gN+gL,2),GD)], 1):
        if v is None:
            ws.cell(row=ri,column=ci).fill = fill(DARK); continue
        c = ws.cell(row=ri, column=ci, value=v)
        c.font = fnt(clr, bold=True, size=10)
        c.fill = fill(DARK); c.alignment = aln("right")
    ws.row_dimensions[ri].height = 26
    ws.freeze_panes = "A5"

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(length))
        zavodlar = body.get("zavodlar", [])
        dan = body.get("dan")
        gacha = body.get("gacha")
        filter_zavod = body.get("zavod")

        if dan and gacha:
            d1 = ".".join(reversed(dan.split("-")))
            d2 = ".".join(reversed(gacha.split("-")))
            davr_label = d1 + " - " + d2
        else:
            davr_label = "Hammasi"

        wb = Workbook()
        if filter_zavod:
            ws1 = wb.active; ws1.title = filter_zavod[:30]
            make_kc_sheet(ws1, zavodlar, filter_zavod, dan, gacha, davr_label)
            ws2 = wb.create_sheet("Hisobot")
            make_hisobot_sheet(ws2, zavodlar, filter_zavod, dan, gacha, davr_label)
        else:
            ws1 = wb.active; ws1.title = "Umumiy"
            make_kc_sheet(ws1, zavodlar, None, dan, gacha, davr_label)
            for z in zavodlar:
                wsz = wb.create_sheet(z["nom"][:30])
                make_kc_sheet(wsz, zavodlar, z["nom"], dan, gacha, davr_label)
            wsh = wb.create_sheet("Hisobot")
            make_hisobot_sheet(wsh, zavodlar, None, dan, gacha, davr_label)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=tilla-hisobot.xlsx")
        self.send_header("Content-Length", str(len(xlsx_bytes)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
